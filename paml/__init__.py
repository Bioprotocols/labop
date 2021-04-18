from sbol_factory import SBOLFactory, Document, ValidationReport, UMLFactory
import sbol3
import os
import posixpath
import openpyxl
import numpy

# Import ontology
__factory__ = SBOLFactory(locals(), 
                          posixpath.join(os.path.dirname(os.path.realpath(__file__)),
                                         'paml.ttl'),
                          'http://bioprotocols.org/paml#')
__umlfactory__ = UMLFactory(__factory__)

#########################################
# Kludge for getting parents and toplevels
# TODO: remove after resolution of https://github.com/SynBioDex/pySBOL3/issues/234
def identified_get_parent(self):
    if self.identity:
        return self.document.find(self.identity.rsplit('/',1)[0])
    else:
        return None
sbol3.Identified.get_parent = identified_get_parent

def identified_get_toplevel(self):
    if isinstance(self, sbol3.TopLevel):
        return self
    else:
        parent = self.get_parent()
        if parent:
            return identified_get_toplevel(parent)
        else:
            return None
sbol3.Identified.get_toplevel = identified_get_toplevel


###########################################
# Define extension methods for Activity
def activity_input_flows(self):
    return {x for x in self.get_toplevel().flows if
            (x.sink.lookup() == self) or
            (isinstance(self, Executable) and x.sink.lookup() in self.input)}
Activity.input_flows = activity_input_flows

def activity_output_flows(self):
    return {x for x in self.get_toplevel().flows if
            (x.source.lookup() == self) or
            (isinstance(self, Executable) and x.source.lookup() in self.output)}
Activity.output_flows = activity_output_flows

def activity_direct_input_flows(self):
    return {x for x in self.get_toplevel().flows if (x.sink.lookup() == self)}
Activity.direct_input_flows = activity_direct_input_flows

def activity_direct_output_flows(self):
    return {x for x in self.get_toplevel().flows if (x.source.lookup() == self)}
Activity.direct_output_flows = activity_direct_output_flows

###########################################
# Define extension methods for Primitive
def primitive_add_input(self, name, type, optional=False):
    pin = PrimitivePinSpecification(name = name, type = type)
    pin.optional = optional
    self.input.append(pin)
    return pin
# Monkey patch
Primitive.add_input = primitive_add_input

def primitive_add_output(self, name, type):
    pin = PrimitivePinSpecification(name = name, type = type)
    self.output.append(pin)
    return pin
# Monkey patch
Primitive.add_output = primitive_add_output


###########################################
# Define extension methods for Executable

# Executable factory functions (can't override initialization)
def executable_make_pins(self, specification, **input_pin_map):
    # first, make sure that all of the keyword arguments are in the inputs of the selected primitive
    unmatched_keys = [key for key in input_pin_map.keys() if key not in (i.name for i in specification.input)]
    assert not unmatched_keys, ValueError("Specification for '"+specification.display_id+"' does not have inputs: "+str(unmatched_keys))

    # Instantiate input pins
    for pin_spec in specification.input:
        if pin_spec.name not in input_pin_map: # Note: affected by pySBOL3 issue 229
            val = None # if not wired to a constant, it will be a generic Pin
        else:
            val = input_pin_map[pin_spec.name]
            # Turning off type checking for now, since I don't know how to check subclass relations efficiently
            #if val.type_uri != pin_spec.type:
            #    raise TypeError(f'Value for input pin "{pin_spec.name or pin_spec.display_id}" must be of type {pin_spec.type}')

        # Construct Pin of appropriate subtype
        if val:
            if isinstance(val, sbol3.TopLevel) or isinstance(val, Location):
                pin = ReferenceValuePin()
            elif isinstance(val, sbol3.Identified):
                pin = LocalValuePin()
            else:
                pin = SimpleValuePin()
            pin.value = val
        else:
            pin = Pin()

        # Set properties
        pin.instance_of = pin_spec
        pin.name = pin_spec.name
        if pin_spec.type:
            pin.type = pin_spec.type
        self.input.append(pin)

    # Instantiate output pins
    for pin_spec in specification.output:
        # Construct Pin of appropriate subtype
        pin = Pin()

        # Set properties
        pin.name = pin_spec.name
        pin.instance_of = pin_spec
        self.output.append(pin)
# monkey patch
Executable.make_pins = executable_make_pins

def Executable_input_pin(self, pin_name):
    #pin_set = [x for x in self.input if x.instance_of.lookup().name == pin_name or doc.find(x.instance_of).display_id == pin_name] # note: pySBOL3 issue 229
    pin_set = [x for x in self.input if self.document.find(x.instance_of).name == pin_name or self.document.find(x.instance_of).display_id == pin_name] # workaround for PAML issue #5
    assert len(pin_set)>0, ValueError("Couldn't find input pin named "+pin_name)
    assert len(pin_set)<2, ValueError("Found more than one input pin named "+pin_name)
    return pin_set[0]
Executable.input_pin = Executable_input_pin

def Executable_output_pin(self, pin_name):
    #pin_set = [x for x in self.output if x.instance_of.lookup().name == pin_name or doc.find(x.instance_of).display_id == pin_name] # note: pySBOL3 issue 229
    pin_set = [x for x in self.output if self.document.find(x.instance_of).name == pin_name or self.document.find(x.instance_of).display_id == pin_name] # workaround for PAML issue #5
    assert len(pin_set)>0, ValueError("Couldn't find output pin named "+pin_name)
    assert len(pin_set)<2, ValueError("Found more than one output pin named "+pin_name)
    return pin_set[0]
Executable.output_pin = Executable_output_pin


def make_PrimitiveExecutable(primitive: Primitive, **input_pin_map):
    self = PrimitiveExecutable(instance_of = primitive)
    self.make_pins(primitive, **input_pin_map)
    return self


def make_SubProtocol(protocol: Protocol, **input_pin_map):
    self = SubProtocol(instance_of = protocol)
    self.make_pins(protocol, **input_pin_map)
    return self


# Get the Value activity associated with the specified input of a subprotocol
def subprotocol_input_value(self: SubProtocol, pin: Pin):
    assert pin in self.input, ValueError("SubProtocol '"+self.identity+"' does not have an input Pin '"+pin+"'")
    return pin.instance_of.lookup().activity
SubProtocol.input_value = subprotocol_input_value

# Get the Value activity associated with the specified output of a subprotocol
def subprotocol_output_value(self: SubProtocol, pin: Pin):
    assert pin in self.output, ValueError("SubProtocol '" + self.identity + "' does not have an output Pin '" + pin + "'")
    return pin.instance_of.lookup().activity
SubProtocol.output_value = subprotocol_output_value


###########################################
# Define extension methods for Protocol


def protocol_contains_activity(self, activity):
    return (activity in self.activities) or \
           next((x for x in self.activities if
                 isinstance(x, Executable) and (activity in x.input or activity in x.output)), False)
# Monkey patch:
Protocol.contains_activity = protocol_contains_activity


def Protocol_initial(self):
    initial = [a for a in self.activities if isinstance(a, Initial)]
    if not initial:
        self.activities.append(Initial())
        return Protocol_initial(self)
    elif len(initial)==1:
        return initial[0]
    else:
        raise ValueError("Protocol '"+self.display_id+"'must have only one Initial node, but found "+str(len(initial)))
# Monkey patch:
Protocol.initial = Protocol_initial


def Protocol_final(self):
    final = [a for a in self.activities if isinstance(a, Final)]
    if not final:
        self.activities.append(Final())
        return Protocol_final(self)
    elif len(final)==1:
        return final[0]
    else:
        raise ValueError("Protocol '"+self.display_id+"'must have only one Final node, but found "+str(len(final)))
# Monkey patch:
Protocol.final = Protocol_final


def protocol_add_input(self, name, **kwargs):
    input = Value(name=name, **kwargs)
    self.activities.append(input)
    input_spec = ProtocolPinSpecification(name=name, activity = input)
    self.input.append(input_spec)
    return input
# Monkey patch:
Protocol.add_input = protocol_add_input


def protocol_add_output(self, name, value_source:Activity=None):
    output = Value()
    self.activities.append(output)
    output_spec = ProtocolPinSpecification(name=name, activity = output)
    self.output.append(output_spec)
    if value_source:
        self.add_flow(value_source, output)
    return output
# Monkey patch:
Protocol.add_output = protocol_add_output


# Create and add an execution of a primitive to a protocol
def protocol_execute_primitive(self, primitive: Primitive, **input_pin_map):
    if isinstance(primitive,str):
        primitive = get_primitive(self.document, primitive)
    # strip any activities in the pin map, which will be held for connecting via flows instead
    activity_inputs = {k: v for k, v in input_pin_map.items() if isinstance(v,Activity)}
    non_activity_inputs = {k: v for k, v in input_pin_map.items() if k not in activity_inputs}
    pe = make_PrimitiveExecutable(primitive, **non_activity_inputs)
    self.activities.append(pe)
    # add flows for activities being connected implicitly
    for k,v in activity_inputs.items():
        self.add_flow(v, pe.input_pin(k))
    return pe
# Monkey patch:
Protocol.execute_primitive = protocol_execute_primitive


# Create and add an execution of a subprotocol to a protocol
def protocol_execute_subprotocol(self, protocol: Protocol, **input_pin_map):
    # strip any activities in the pin map, which will be held for connecting via flows instead
    activity_inputs = {k: v for k, v in input_pin_map.items() if isinstance(v,Activity)}
    non_activity_inputs = {k: v for k, v in input_pin_map.items() if k not in activity_inputs}
    sub = make_SubProtocol(protocol, **non_activity_inputs)
    self.activities.append(sub)
    # add flows for activities being connected implicitly
    for k,v in activity_inputs.items():
        self.add_flow(v, sub.input_pin(k))
    return sub
# Monkey patch:
Protocol.execute_subprotocol = protocol_execute_subprotocol


# Create and add a flow between the designated child source and sink activities
def protocol_add_flow(self, source, sink):
    assert self.contains_activity(source), ValueError(
        'Source activity ' + source.identity + ' is not a member of protocol ' + self.identity)
    assert self.contains_activity(sink), ValueError(
        'Sink activity ' + sink.identity + ' is not a member of protocol ' + self.identity)
    flow = Flow(source = source, sink = sink)
    self.flows.append(flow)
    return flow
# Monkey patch:
Protocol.add_flow = protocol_add_flow


###########################################
# Define extension methods for LocatedSamples

# Transform an Excel-style range (col:row, inclusive, alpha-numeric) to numpy-style (row:col, start/stop, numeric)
def excel_to_numpy_range(excel_range):
    bounds = openpyxl.utils.cell.range_boundaries(excel_range)
    return [bounds[1]-1,bounds[0]-1,bounds[3],bounds[2]]

def numpy_to_excel_range(top,left,bottom,right):
    if top+1==bottom and left+1==right: # degenerate case of a single cell
        return openpyxl.utils.cell.get_column_letter(left+1)+str(top+1)
    else:
        return openpyxl.utils.cell.get_column_letter(left+1)+str(top+1) + ":" + \
               openpyxl.utils.cell.get_column_letter(right) + str(bottom)

def extract_range_from_top_left(region: numpy.ndarray):
    # find the largest rectangular region starting at the first top-left zero
    top = numpy.where(region)[0][0]
    left = numpy.where(region)[1][0]
    right = numpy.where(region[top,:])[0][-1]+1
    for bottom in range(top,region.shape[0]):
        if not region[bottom,left:right].all():
            bottom -= 1
            break
    bottom += 1 # adjust to stop coordinate
    region[top:bottom, left:right] = False
    return numpy_to_excel_range(top, left, bottom, right)

def reduce_range_set(ranges):
    assert len(ranges)>0, "Range set to reduce must have at least one element"
    bounds = [max(openpyxl.utils.cell.range_boundaries(r)[i] for r in ranges) for i in range(2,4)]
    region = numpy.zeros([bounds[1],bounds[0]],dtype=bool) # make an array of zeros
    # mark each range in turn, ensuring that they don't overlap
    for r in ranges:
        nr = excel_to_numpy_range(r)
        assert not (region[nr[0]:nr[2], nr[1]:nr[3]]).any(), ValueError("Found overlapping range in "+str(ranges))
        region[nr[0]:nr[2], nr[1]:nr[3]] = True
    # pull chunks out until all zeros
    reduced = set()
    while region.any():
        reduced.add(extract_range_from_top_left(region))
    return reduced


def heterogeneous_samples_locations(self: HeterogeneousSamples):
    return [s.in_location for s in self.replicate_samples for l in s.in_location]
HeterogeneousSamples.locations = heterogeneous_samples_locations

def heterogeneous_samples_containers(self : HeterogeneousSamples):
    return {(l if isinstance(l, Container) else l.in_container) for l in self.locations()} # pull containers from locations
HeterogeneousSamples.containers = heterogeneous_samples_containers

def heterogeneous_samples_locations_by_container(self : HeterogeneousSamples):
    locations = self.locations()
    containers = self.containers()
    return {c:{l for l in locations if l==c or (isinstance(l, ContainerCoordinates) and l.in_container==c)} for c in containers}
HeterogeneousSamples.locations_by_container = heterogeneous_samples_locations_by_container


#########################################
# Library handling
loaded_libraries = {}

def import_library(library:str, file_format:str = 'ttl', nickname:str=None ):
    if not nickname:
        nickname = library
    if not os.path.isfile(library):
        library = posixpath.join(os.path.dirname(os.path.realpath(__file__)),
                                 ('lib/'+library+'.ttl'))
    # read in the library and put the document in the library collection
    lib = sbol3.Document()
    lib.read(library, file_format)
    loaded_libraries[nickname] = lib

def get_primitive(doc:sbol3.Document, name:str):
    found = doc.find(name)
    if not found:
        found = {n:l.find(name) for (n,l) in loaded_libraries.items() if l.find(name)}
        assert len(found)<2, ValueError("Ambiguous primitive: found '"+name+"' in multiple libraries: "+str(found.keys()))
        assert len(found)>0, ValueError("Couldn't find primitive '"+name+"' in any library")
        found = next(iter(found.values())).copy(doc)
    return found

