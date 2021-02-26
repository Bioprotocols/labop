import sbol3
import paml
import tyto
import openpyxl
from copy import copy

def markdown_measure(measure):
    return str(measure.value) + ' ' + str(tyto.OM.get_term_by_uri(measure.unit))

def markdown_component(component):
    return '[' + (component.display_id if (component.name is None) else component.name) + ']('+component.types[0]+')'

def markdown_container(container):
    return '[' + (container.display_id if (container.name is None) else container.name) + ']('+container.type+')'

def markdown_containercoodinates(coordinates):
    return markdown_container(coordinates.inContainer.lookup()) + ' ' + coordinates.coordinates

def markdown_location(location):
    if isinstance(location, paml.ContainerCoordinates):
        return markdown_containercoodinates(location)
    elif isinstance(location, paml.Container):
        return markdown_container(location)
    else:
        return str(location)

def markdown_mergedlocations(location_list):
    this = markdown_location(location_list.pop())
    if len(location_list) == 0:
        return this
    if len(location_list) == 1:
        return this + ' and ' + markdown_mergedlocations(location_list)
    else:
        return this + ', ' + markdown_mergedlocations(location_list)

def markdown_flow_value(document, value):
    if isinstance(value, paml.ReplicateSamples):
        return markdown_mergedlocations({x.lookup() for x in value.inLocation})
    elif isinstance(value, paml.HeterogeneousSamples):
        return markdown_mergedlocations({document.find(loc) for rep in value.hasReplicateSamples for loc in rep.inLocation})
    # if we fall through to here:
    return str(value)


############
# BUG: this should not need the document; this is due to pySBOL3 bug #176
def input_pin_value(document, executable, pin_name):
    pin_set = [x for x in executable.input if document.find(x.instanceOf).name == pin_name]
    if len(pin_set) != 1:
        return "[couldn't find input "+pin_name+"]"
    pin = pin_set[0]
    if isinstance(pin, paml.LocalValuePin):
        if isinstance(pin.value, paml.Measure):
            return markdown_measure(pin.value)
    elif isinstance(pin, paml.ReferenceValuePin):
        value = pin.value.lookup()
        if isinstance(value, sbol3.Component):
            return markdown_component(value)
        elif isinstance(value, paml.Location):
            return markdown_location(value)
    elif flow_values[next(x for x in protocol.hasFlow if x.sink.lookup() in executable.input)]:
        value = flow_values[next(x for x in protocol.hasFlow if x.sink.lookup() in executable.input)]
        return markdown_flow_value(document, value)
    # if we fall through to here:
    return str(pin)

############
# BUG: this should not need the document; this is due to pySBOL3 bug #176
def markdown_provision(document, executable):
    volume = input_pin_value(document, executable, 'volume')
    resource = input_pin_value(document, executable, 'resource')
    location = input_pin_value(document, executable, 'location')
    instruction = 'Pipette '+volume+' of '+resource+' into '+location+'\n'
    return instruction

############
# BUG: this should not need the document; this is due to pySBOL3 bug #176
def markdown_absorbance(document, executable):
    location = input_pin_value(document, executable, 'location')
    wavelength = input_pin_value(document, executable, 'wavelength')
    instruction = 'Measure absorbance of '+location+' at '+wavelength+'\n'
    return instruction

#################
# All this stuff should be transformed into visitor patterns
primitive_library = {
    'https://bioprotocols.org/paml/primitives/Provision' : markdown_provision,
    'https://bioprotocols.org/paml/primitives/MeasureAbsorbance' : markdown_absorbance
}

def get_value_flow_input(protocol, activity):
    flows = (x for x in protocol.hasFlow if (x.sink.lookup() == activity)) # need to generalize to multiple
    ########
    # TODO: remove this evil kludge where we're dipping into a global variable
    return flow_values[next(flows)]

def markdown_value(document, protocol, activity):
    return 'Report values from '+markdown_flow_value(document, get_value_flow_input(protocol, activity))+'\n'

def markdown_header(protocol):
    header = '# ' + (protocol.display_id if (protocol.name is None) else protocol.name) + '\n'
    header += '\n'
    header += '## Description:\n' + ('No description given' if protocol.description is None else protocol.description) + '\n'
    return header

#############
# BUG: this should not need the document; this is due to pySBOL3 bug #176
def markdown_activity(document, activity):
    if isinstance(activity, paml.PrimitiveExecutable):
        stepwriter = primitive_library[document.find(activity.instanceOf).identity]
        assert stepwriter
        return stepwriter(document, activity)
    elif isinstance(activity, paml.Value):
        return markdown_value(document, protocol, activity)
    else:
        raise ValueError("Don't know how to serialize activity "+activity)

def markdown_material(component):
    bullet = '* ' + markdown_component(component)
    if component.description is not None: bullet += ': ' + component.description
    bullet += '\n'
    return bullet

def markdown_container_toplevel(container):
    bullet = '* ' + markdown_container(container)
    if container.description is not None: bullet += ': ' + container.description
    return bullet

def unpin_activity(protocol, activity):
    if isinstance(activity,paml.Pin):
        owner = next(x for x in protocol.hasActivity if isinstance(x,paml.Executable) and
                     (activity in x.output or activity in protocol.input))
        return owner
    else:
        return activity

##############################
# Visitors for computing flow types
def get_input_pin(executable, pin_name):
    return next(x for x in executable.input if x.instanceOf.lookup().name == pin_name)
def get_output_pin(executable, pin_name):
    return next(x for x in executable.output if x.instanceOf.lookup().name == pin_name)

def type_from_pin_or_flow(protocol, executable, pin_name, flow_values):
    pin = get_input_pin(executable, pin_name)
    if isinstance(pin, paml.LocalValuePin) or isinstance(pin, paml.ReferenceValuePin):
        return pin.value
    else:
        return flow_values[next(x for x in protocol.hasFlow if (x.sink.lookup() == pin))]

def inference_provision(protocol, executable, flow_values):
    resource = type_from_pin_or_flow(protocol, executable, 'resource', flow_values)
    location = type_from_pin_or_flow(protocol, executable, 'location', flow_values)
    samples = paml.ReplicateSamples()
    samples.inLocation.append(location)
    samples.specification = resource
    samples_flow = next(x for x in protocol.hasFlow if x.source.lookup()==get_output_pin(executable, 'samples'))
    return {samples_flow : samples}

def inference_absorbance(protocol, executable, flow_values):
    location = type_from_pin_or_flow(protocol, executable, 'location', flow_values)
    # TODO: make this a LocatedData rather than just copying the samples
    # samples = paml.LocatedData()
    samples_flow = next(x for x in protocol.hasFlow if x.source.lookup()==get_output_pin(executable, 'measurements'))
    return {samples_flow : location}

primitive_inference = {
    'https://bioprotocols.org/paml/primitives/Provision' : inference_provision,
    'https://bioprotocols.org/paml/primitives/MeasureAbsorbance' : inference_absorbance
}

def primitive_types(protocol, activity, flow_values):
    inference_function = primitive_inference[activity.instanceOf.lookup().identity]
    assert inference_function
    return inference_function(protocol, activity, flow_values)

def inflows_satisfied(protocol,flow_values,activity):
    inflows = {x for x in protocol.hasFlow if (x.sink.lookup() == activity) or (isinstance(activity,paml.Executable) and x.sink.lookup() in activity.input)}
    unsatisfied = inflows - flow_values.keys()
    return len(unsatisfied) == 0

# kludge for now
def join_locations(value_set):
    if not value_set:
        return paml.HeterogeneousSamples()
    next = value_set.pop()
    rest = join_locations(value_set)
    if isinstance(next, paml.ReplicateSamples):
        rest.hasReplicateSamples.append(next)
    elif isinstance(next, paml.HeterogeneousSamples):
        for x in next.hasReplicateSamples: rest.hasReplicateSamples.append(x)
    else:
        raise ValueError("Don't know how to join locations for "+str(value_set))
    return rest

def join_values(value_set):
    if all(isinstance(x,paml.LocatedSamples) for x in value_set):
        return join_locations(value_set)
    # if we fall through to the end, then we didn't know how to infer
    raise ValueError("Don't know how to join values types for "+str(value_set))

# returns a dictionary of outflow to type
def outflow_types(protocol,activity,flow_values):
    direct_outflows = {x for x in protocol.hasFlow if (x.source.lookup() == activity)}
    direct_inflows = {x for x in protocol.hasFlow if (x.sink.lookup() == activity)}
    if isinstance(activity, paml.Control):
        if isinstance(activity, paml.Initial):
            return {x: None for x in direct_outflows}
        elif isinstance(activity, paml.Final):
            assert len(direct_outflows)==0 # should be no outputs
            return {}
        elif isinstance(activity, paml.Fork) or isinstance(activity, paml.Decision):
            assert len(direct_inflows)==1 # should be precisely one input
            in_type = flow_values[next(direct_inflows)]
            return {x: in_type for x in direct_outflows}
        elif isinstance(activity, paml.Join):
            assert len(direct_outflows)==1 # should be precisely one output
            return {direct_outflows.pop() : join_values({flow_values[x] for x in direct_inflows})}
    elif isinstance(activity, paml.PrimitiveExecutable):
        direct_types = {x:None for x in direct_outflows}
        pin_types = primitive_types(protocol, activity, flow_values)
        return direct_types | pin_types
    elif isinstance(activity, paml.Value):
        assert len(direct_outflows) == 1  # should be precisely one output
        return {x: None for x in direct_outflows}
    # if we fall through to the end, then we didn't know how to infer
    raise ValueError("Don't know how to infer outflow types for "+str(activity))



##############################
# For serializing activities
def direct_precedents(protocol, activity):
    assert activity in protocol.hasActivity
    flows = (x for x in protocol.hasFlow if (x.sink.lookup() == activity) or (isinstance(activity,paml.Executable) and x.sink.lookup() in activity.input))
    precedents = {unpin_activity(protocol,x.source.lookup()) for x in flows}
    return precedents

##############################
# Get the protocol
print('Reading document')

doc = paml.Document()
doc.read('igem_ludox_draft.json','json-ld')

# extract set of protocols from document
protocols = {x for x in doc.objects if isinstance(x, paml.Protocol)}
if len(protocols)==0:
    raise ValueError("Cannot find any protocols in document")
elif len(protocols)>1:
    raise ValueError("Found multiple protocols; don't know which to write")

# pull the first and only
protocol = protocols.pop()

print('Found protocol: '+protocol.display_id)

##############################
# Infer values carried on flows
print('Inferring flow values')

flow_values = {} # dictionary of flow : type mappings
pending_activities = set(protocol.hasActivity)
while pending_activities:
    non_blocked = {x for x in pending_activities if inflows_satisfied(protocol,flow_values,x)}
    if not non_blocked:
        raise ValueError("Could not infer all flow types: circular dependency?")
        break
    for activity in non_blocked:
        flow_values.update(outflow_types(protocol,activity,flow_values))
    pending_activities -= non_blocked


##############################
# Serialize order of steps
print('Serializing activities')

serialized_activities = []
pending_activities = set(protocol.hasActivity)
while pending_activities:
    non_blocked = {x for x in pending_activities if not (direct_precedents(protocol,x) & set(pending_activities))}
    if not non_blocked:
        raise ValueError("Could not serialize all activities: circular dependency?")
        break
    serialized_activities += non_blocked
    pending_activities -= non_blocked

assert isinstance(serialized_activities[0],paml.Initial)
assert isinstance(serialized_activities[-1],paml.Final)

# filter out control flow statements
serialized_noncontrol_activities = [x for x in serialized_activities if not isinstance(x,paml.Control)]

##############################
# Write to a markdown file

print('Writing markdown file')

with open(protocol.display_id+'.md', 'w') as file:
    file.write(markdown_header(protocol))

    file.write('\n\n ## Materials\n')
    for material in protocol.material:
        file.write(markdown_material(material.lookup()))
    for container in (x for x in protocol.hasLocation if isinstance(x,paml.Container)):
        file.write(markdown_container_toplevel(container))

    file.write('\n\n ## Steps\n')
    for step in range(len(serialized_noncontrol_activities)):
        file.write('### Step '+str(step+1)+'\n'+markdown_activity(doc,serialized_noncontrol_activities[step])+'\n')

    # make sure the file is fully written
    file.flush()

##############################
# Write to an accompanying Excel file

def excel_container_name(container):
   return (container.display_id if (container.name is None) else container.name)

# def excel_write_container(ws, row_offset, container):
#     return '[' + (container.display_id if (container.name is None) else container.name) + ']('+container.type+')'

def excel_write_containercoodinates(ws, row_offset, col_offset, coordinates):
    # get the column letter
    col = openpyxl.utils.cell.get_column_letter(col_offset+1)
    # make the header
    ws[col+str(row_offset)] = excel_container_name(coordinates.inContainer.lookup())
    # write the materials
    block = openpyxl.utils.cell.range_boundaries(coordinates.coordinates)
    # use plate coordinates, which are the opposite of excel coordinates
    height = block[2]-block[0]+1
    width = block[3]-block[1]+1
    # write the plate coordinate frame
    for plate_row in range(height):
        coord = col+str(row_offset+plate_row+2)
        ws[coord] = openpyxl.utils.cell.get_column_letter(block[0]+plate_row)
        ws[coord].fill = copy(fixed_style)
    for plate_col in range(width):
        coord = openpyxl.utils.cell.get_column_letter(col_offset+plate_col+2)+str(row_offset+1)
        ws[coord] = str(block[1]+plate_col)
        ws[coord].fill = copy(fixed_style)
    # style the blanks
    for plate_row in range(height):
        for plate_col in range(width):
            coord = openpyxl.utils.cell.get_column_letter(col_offset+plate_col+2)+str(row_offset+plate_row+2)
            ws[coord].fill = copy(entry_style)
            ws[coord].alignment = openpyxl.styles.Alignment(horizontal="center")

    return (height+2, width+1)

def excel_write_location(ws, row_offset, col_offset, location):
    if isinstance(location, paml.ContainerCoordinates):
        return excel_write_containercoodinates(ws, row_offset, col_offset, location)
    # elif isinstance(location, paml.Container):
    #     return excel_write_container(location)
    else:
        return str(location)

def excel_write_mergedlocations(ws, row_offset, location_list):
    col_offset = 0
    block_height = 0
    while location_list:
        block = excel_write_location(ws, row_offset, col_offset, location_list.pop())
        col_offset += block[1] + 1
        block_height = max(block_height,block[0])
    return block_height

def excel_write_flow_value(document, value, ws, row_offset):
    if isinstance(value, paml.ReplicateSamples):
        return excel_write_mergedlocations(ws, row_offset, {x.lookup() for x in value.inLocation})
    elif isinstance(value, paml.HeterogeneousSamples):
        return excel_write_mergedlocations(ws, row_offset, {document.find(loc) for rep in value.hasReplicateSamples for loc in rep.inLocation})
    # if we fall through to here:
    return str(value)

print('Writing Excel file')

wb = openpyxl.load_workbook(filename = 'template.xlsx')
ws = wb.active # get the default worksheet

# write header & metadata
ws.title = "Data Reporting"
ws['D1'] = protocol.name
fixed_style = ws['A2'].fill
entry_style = ws['C2'].fill
header_style = ws['A1'].font
row_offset = 7 # starting point for entries

# write each value set, incrementing each time
value_steps = (step for step in range(len(serialized_noncontrol_activities)) if isinstance(serialized_noncontrol_activities[step], paml.Value))
for step in value_steps:
    coord = 'A'+str(row_offset)
    ws[coord] = 'Report from Step '+str(step+1)
    ws[coord].font = copy(header_style)
    value_locations = get_value_flow_input(protocol, serialized_noncontrol_activities[step])
    block_height = excel_write_flow_value(doc, value_locations, ws, row_offset+1)
    row_offset += block_height + 2

wb.save(protocol.display_id+'.xlsx')

print('Export complete')
