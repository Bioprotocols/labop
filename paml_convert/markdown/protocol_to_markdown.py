import sbol3
import paml
import uml
import paml.type_inference
import openpyxl
import numpy
from IPython.display import Markdown


###########################################
# Functions for reasoning about ranges

# Transform an Excel-style range (col:row, inclusive, alpha-numeric) to numpy-style (row:col, start/stop, numeric)
def excel_to_numpy_range(excel_range):
    bounds = openpyxl.utils.cell.range_boundaries(excel_range)
    return [bounds[1]-1,bounds[0]-1,bounds[3],bounds[2]]

def numpy_to_excel_range(top,left,bottom,right):
    if top+1==bottom and left+1==right: # degenerate case of a single cell
        return openpyxl.utils.cell.get_column_letter(left+1)+str(top+1)
    else:
        return openpyxl.utils.cell.get_column_letter(left+1)+str(top+1) + ":" + \
               openpyxl.utils.cell.get_column_letter(right) + str(bottom)

def extract_range_from_top_left(region: numpy.ndarray):
    # find the largest rectangular region starting at the first top-left zero
    top = numpy.where(region)[0][0]
    left = numpy.where(region)[1][0]
    right = numpy.where(region[top,:])[0][-1]+1
    for bottom in range(top,region.shape[0]):
        if not region[bottom,left:right].all():
            bottom -= 1
            break
    bottom += 1 # adjust to stop coordinate
    region[top:bottom, left:right] = False
    return numpy_to_excel_range(top, left, bottom, right)

def reduce_range_set(ranges):
    assert len(ranges)>0, "Range set to reduce must have at least one element"
    bounds = [max(openpyxl.utils.cell.range_boundaries(r)[i] for r in ranges) for i in range(2,4)]
    region = numpy.zeros([bounds[1],bounds[0]],dtype=bool) # make an array of zeros
    # mark each range in turn, ensuring that they don't overlap
    for r in ranges:
        nr = excel_to_numpy_range(r)
        assert not (region[nr[0]:nr[2], nr[1]:nr[3]]).any(), ValueError("Found overlapping range in "+str(ranges))
        region[nr[0]:nr[2], nr[1]:nr[3]] = True
    # pull chunks out until all zeros
    reduced = set()
    while region.any():
        reduced.add(extract_range_from_top_left(region))
    return reduced


##############################################
# Converter state object, to be carried along with "to_markdown" functions
#
# TODO: make this build PROV-O as it goes?
class MarkdownConverter():
    def __init__(self, document: sbol3.Document):
        self.document = document
        #self.protocol_typing = paml.type_inference.ProtocolTyping()

    def markdown_header(self, protocol):
        header = '# ' + (protocol.display_id if (protocol.name is None) else protocol.name) + '\n'
        header += '\n'
        header += '## Description:\n' + (
            'No description given' if protocol.description is None else protocol.description) + '\n'
        return header

    # Entry-point for document conversion
    # TODO: allow us to control the name of the output
    def convert(self, execution, out=None):
        # protocol argument can be either string, URI, or paml.Protocol
        if not isinstance(execution, paml.ProtocolExecution):
            execution = self.document.find(execution)

        # print('Inferring flow values')
        # self.protocol_typing.infer_typing(protocol)

        print('Serializing activities')
        serialized_noncontrol_activities = serialize_activities(execution)

        print('Writing markdown file')
        markdown = write_markdown_file(execution, serialized_noncontrol_activities, self, out=out)

        print('Writing Excel file')
        # write_excel_file(protocol, serialized_noncontrol_activities, self)

        print('Export complete')
        return markdown



# ##############################################
# # Direct conversion of individual objects to their markdown representations
#
# def list_to_markdown(l: list, mdc: MarkdownConverter):
#     if len(l) == 0:
#         return '' ## TODO: kludge: this shouldn't be happening
#     this = l.pop().to_markdown(mdc)
#     if len(l) == 0:
#         return this
#     if len(l) == 1:
#         return this + ' and ' + list_to_markdown(l, mdc)
#     else:
#         return this + ', ' + list_to_markdown(l, mdc)
#
# def strlist_to_markdown(l: list, mdc: MarkdownConverter):
#     if len(l) == 0:
#         return '' ## TODO: kludge: this shouldn't be happening
#     this = l.pop()
#     if len(l) == 0:
#         return this
#     if len(l) == 1:
#         return this + ' and ' + strlist_to_markdown(l, mdc)
#     else:
#         return this + ', ' + strlist_to_markdown(l, mdc)
#
# def measure_to_markdown(self: sbol3.Measure, mdc: MarkdownConverter):
#     # TODO: convert large numbers to friendlier units
#     unit = (mdc.document.find(self.unit).name if mdc.document.find(self.unit) else tyto.OM.get_term_by_uri(self.unit))
#     if unit=="number":  # special case: don't need to say unit for pure numbers
#         unit = ''
#     return str(self.value) + ' ' + str(unit)
# sbol3.Measure.to_markdown = measure_to_markdown
#
#
# def component_to_markdown(self: sbol3.Component, mdc: MarkdownConverter):
#     return '[' + (self.display_id if (self.name is None) else self.name) + ']('+self.types[0]+')'
# sbol3.Component.to_markdown = component_to_markdown
#
#
# def container_to_markdown(self: paml.Container, mdc: MarkdownConverter):
#     return '[' + (self.display_id if (self.name is None) else self.name) + ']('+self.type+')'
# paml.Container.to_markdown = container_to_markdown
#
#
# def containercoodinates_to_markdown(self: paml.ContainerCoordinates, mdc: MarkdownConverter):
#     return mdc.document.find(self.in_container).to_markdown(mdc) + ' ' + self.coordinates # TODO: figure out how to set document to enable changing doc.find to lookup
# paml.ContainerCoordinates.to_markdown = containercoodinates_to_markdown
#
#
#
# # Helper function for reducing coordinates shown for LocatedSamples
# def markdown_mergedlocations(location_list, mdc: MarkdownConverter):
#     containers = [c for c in location_list if isinstance(c, paml.Container)]
#     coords = {coord for coord in location_list if isinstance(coord, paml.ContainerCoordinates)}
#     reduced = []
#     for c in {coord.in_container.lookup() for coord in coords}:
#         ranges = reduce_range_set({l.coordinates for l in coords if l.in_container.lookup()==c})
#         for r in ranges:  # BUG: should be a list comprehension, but in_container argument can't be set in constructor
#             cc = paml.ContainerCoordinates(coordinates=r)
#             cc.in_container = c
#             reduced.append(cc)
#     return list_to_markdown(containers+reduced, mdc)
#
#
# def locateddata_to_markdown(self: paml.LocatedData, mdc: MarkdownConverter):
#     return self.from_samples.to_markdown(mdc)
# paml.LocatedData.to_markdown = locateddata_to_markdown
#
# def locatedsamples_to_markdown(self: paml.ReplicateSamples, _: MarkdownConverter):
#     return self.name  # TODO: can we do better than this kludge?
# paml.LocatedSamples.to_markdown = locatedsamples_to_markdown
#
#
# def replicatesamples_to_markdown(self: paml.ReplicateSamples, mdc: MarkdownConverter):
#     return markdown_mergedlocations({mdc.document.find(x) for x in self.in_location}, mdc)
# paml.ReplicateSamples.to_markdown = replicatesamples_to_markdown
#
#
# def heterogeneoussamples_to_markdown(self: paml.HeterogeneousSamples, mdc: MarkdownConverter):
#     return markdown_mergedlocations({mdc.document.find(loc) for rep in self.replicate_samples for loc in rep.in_location}, mdc)
# paml.HeterogeneousSamples.to_markdown = heterogeneoussamples_to_markdown
#
#
# def integerconstantpin_to_markdown(self: paml.IntegerConstantPin, mdc: MarkdownConverter):
#     return str(self.value)
# paml.IntegerConstantPin.to_markdown = integerconstantpin_to_markdown
#
# def stringconstantpin_to_markdown(self: paml.StringConstantPin, mdc: MarkdownConverter):
#     return str(self.value)
# paml.StringConstantPin.to_markdown = stringconstantpin_to_markdown
#
#
# def localvaluepin_to_markdown(self: paml.LocalValuePin, mdc: MarkdownConverter):
#     return self.value.to_markdown(mdc)
# paml.LocalValuePin.to_markdown = localvaluepin_to_markdown
#
#
# def referencevaluepin_to_markdown(self: paml.ReferenceValuePin, mdc: MarkdownConverter):
#     return self.value.lookup().to_markdown(mdc)
# paml.ReferenceValuePin.to_markdown = referencevaluepin_to_markdown
#
# # A non-constant pin needs to pull its value from the flow
# def pin_to_markdown(self: paml.Pin, mdc: MarkdownConverter):
#     inflows = self.input_flows()
#     assert len(inflows)==1, ValueError('Pin has more than one input flow: '+self.identity)
#     value = mdc.protocol_typing.flow_values[inflows.pop()]
#     return value.to_markdown(mdc)
# paml.Pin.to_markdown = pin_to_markdown
#
# ###############################
# # Base activities to markdown
#
# def primitiveexecutable_to_markdown(self: paml.PrimitiveExecutable, mdc: MarkdownConverter):
#     stepwriter = primitive_to_markdown_functions[mdc.document.find(self.instance_of).identity]
#     return stepwriter(self, mdc)
# paml.PrimitiveExecutable.to_markdown = primitiveexecutable_to_markdown
#
# def subcall_variable_to_markdown(pin,mdc):
#     activity = pin.instance_of.lookup().activity.lookup()
#     return pin.to_markdown(mdc) + " for " + (activity.description if activity.description else activity.name)
#
# def subprotocol_to_markdown(self: paml.SubProtocol, mdc: MarkdownConverter):
#     protocol = self.instance_of.lookup()
#     pname = (protocol.display_id if (protocol.name is None) else protocol.name)
#     input_string = strlist_to_markdown([subcall_variable_to_markdown(pin,mdc) for pin in self.input], mdc)
#     return 'Run protocol "'+pname+'" with inputs: '+input_string
# paml.SubProtocol.to_markdown = subprotocol_to_markdown
#
#
# def value_to_markdown(self: paml.Value, mdc: MarkdownConverter):
#     if is_input_value(self):  # This is an input value, used as a variable
#         value = mdc.protocol_typing.flow_values[self.output_flows().pop()]
#         return 'Protocol input: ' + value.to_markdown(mdc)
#     else:  # This is an output value, used for reporting
#         value = mdc.protocol_typing.flow_values[self.input_flows().pop()]
#         return 'Report values from ' + value.to_markdown(mdc) + '\n'
# paml.Value.to_markdown = value_to_markdown
#
#
#############################
# Other sorts of markdown functions

def markdown_input(input : uml.Parameter, mdc: MarkdownConverter):
    bullet = '* ' + str(input)
    if input.description is not None: bullet += ': ' + input.description
    bullet += '\n'
    return bullet


# def markdown_material(component, mdc: MarkdownConverter):
#     bullet = '* ' + component.to_markdown(mdc)
#     if component.description is not None: bullet += ': ' + component.description
#     bullet += '\n'
#     return bullet
#
#
# def markdown_container_toplevel(container, mdc: MarkdownConverter):
#     # TODO: generalize ontology option for type
#     bullet = '* ' + container.to_markdown(mdc) + ' (['+tyto.NCIT.get_term_by_uri(container.type)+']('+container.type+'))'
#     if container.description is not None: bullet += ': ' + container.description
#     bullet += '\n'
#     return bullet
#
#
#
# ##############################
# # Serialize order of steps
#
# def unpin_activity(protocol, activity):
#     return activity.get_parent() if isinstance(activity, paml.Pin) else activity
#
# def is_input_value(x):
#     return isinstance(x,paml.Value) and \
#            not({f for f in x.input_flows() if not isinstance(f.source.lookup(), paml.Initial)})
#
def serialize_activities(execution: paml.ProtocolExecution):
    serialized_activities = []

    for execution in execution.executions:
        if isinstance(execution, paml.CallBehaviorExecution):
            execution_node = execution.node.lookup()
            serialized_activities.append(execution_node)

    # assert isinstance(serialized_activities[0], paml.Initial)
    # assert isinstance(serialized_activities[-1], paml.Final)

    # filter out control flow statements
    serialized_noncontrol_activities = [x for x in serialized_activities if (not isinstance(x, uml.ControlNode)) and (not isinstance(x, uml.ObjectNode))]
    serialized_noncontrol_activities.reverse()
    return serialized_noncontrol_activities

##############################
# Write to a markdown file

def write_markdown_file(execution: paml.ProtocolExecution, serialized_noncontrol_activities, mdc: MarkdownConverter, out=None):
    protocol = execution.protocol.lookup()
    markdown = mdc.markdown_header(protocol)

    markdown += '\n\n## Protocol Inputs:\n'
    for i in protocol.parameters:
        if i.property_value.direction == uml.PARAMETER_IN:
            markdown += markdown_input(i.property_value, mdc)

    markdown += '\n\n## Materials\n'
    # for material in protocol.material:
    #     file.write(markdown_material(material.lookup(), mdc))

    markdown += '\n\n## Containers\n'
    # for container in (x for x in protocol.locations if isinstance(x, paml.Container)):
    #     file.write(markdown_container_toplevel(container, mdc))

    markdown += '\n\n## Steps\n'
    for step in range(len(serialized_noncontrol_activities)):
        print('Writing step '+str(step)+": "+serialized_noncontrol_activities[step].identity)
        #file.write('### Step ' + str(step + 1) + '\n' + serialized_noncontrol_activities[step].to_markdown(mdc) + '\n')
        # file.write(str(step + 1) + '. ' + serialized_noncontrol_activities[step].to_markdown(mdc) + '\n')
        markdown += str(step + 1) + '. ' + serialized_noncontrol_activities[step].behavior + '\n'

    markdown = Markdown(markdown)

    # make sure the file is fully written
    if out:
        with open(out, 'w') as file:
            markdown.filename = out
            file.write(markdown.data)
            file.flush()
            file.close()

    print('Finished writing markdown')
    return markdown

# ##############################
# # Write to an accompanying Excel file
#
# def excel_container_name(container):
#    return (container.display_id if (container.name is None) else container.name)
#
# # def excel_write_container(ws, row_offset, container):
# #     return '[' + (container.display_id if (container.name is None) else container.name) + ']('+container.type+')'
#
# def excel_write_containercoodinates(ws, row_offset, col_offset, coordinates, specification_URI):
#     fixed_style = ws['A2'].fill
#     entry_style = ws['C2'].fill
#
#     # get the column letter
#     col = openpyxl.utils.cell.get_column_letter(col_offset+1)
#     # make the header
#     ws[col+str(row_offset)] = excel_container_name(coordinates.in_container.lookup())
#     # write the materials
#     block = openpyxl.utils.cell.range_boundaries(coordinates.coordinates)
#     # use plate coordinates, which are the opposite of excel coordinates
#     height = block[2]-block[0]+1
#     width = block[3]-block[1]+1
#     # write the plate coordinate frame
#     for plate_row in range(height):
#         coord = col+str(row_offset+plate_row+2)
#         ws[coord] = openpyxl.utils.cell.get_column_letter(block[0]+plate_row)
#         ws[coord].fill = copy(fixed_style)
#     for plate_col in range(width):
#         coord = openpyxl.utils.cell.get_column_letter(col_offset+plate_col+2)+str(row_offset+1)
#         ws[coord] = str(block[1]+plate_col)
#         ws[coord].fill = copy(fixed_style)
#     # style the blanks
#     for plate_row in range(height):
#         for plate_col in range(width):
#             coord = openpyxl.utils.cell.get_column_letter(col_offset+plate_col+2)+str(row_offset+plate_row+2)
#             ws[coord].fill = copy(entry_style)
#             ws[coord].alignment = openpyxl.styles.Alignment(horizontal="center")
#             plate_coord = openpyxl.utils.cell.get_column_letter(block[0]+plate_row)+str(block[1]+plate_col)
#             ws[coord].comment = openpyxl.comments.Comment(coordinates.in_container+"_"+plate_coord+" "+specification_URI, "PAML autogeneration, do not modify", height=24, width=1000)
#             ws[coord].protection = openpyxl.styles.Protection(locked=False)
#
#     return (height+2, width+1)
#
# def excel_write_location(ws, row_offset, col_offset, location, specification_URI):
#     if isinstance(location, paml.ContainerCoordinates):
#         return excel_write_containercoodinates(ws, row_offset, col_offset, location, specification_URI)
#     # elif isinstance(location, paml.Container):
#     #     return excel_write_container(location)
#     else:
#         return str(location)
#
# def excel_write_mergedlocations(ws, row_offset, location_spec_list):
#     col_offset = 0
#     block_height = 0
#     while location_spec_list:
#         pair = location_spec_list.popitem()
#         block = excel_write_location(ws, row_offset, col_offset, pair[0], pair[1])
#         col_offset += block[1] + 1
#         block_height = max(block_height,block[0])
#     return block_height
#
# # TODO: consolidate the Excel reporting squares, like we've already done for the protocol above
# def excel_write_flow_value(document, value, ws, row_offset):
#     if isinstance(value, paml.LocatedData):
#         value = value.from_samples  # unwrap value
#     if isinstance(value, paml.ReplicateSamples):
#         return excel_write_mergedlocations(ws, row_offset, {x.lookup():value.specification for x in value.in_location})
#     elif isinstance(value, paml.HeterogeneousSamples):
#         return excel_write_mergedlocations(ws, row_offset, {document.find(loc):rep.specification for rep in value.replicate_samples for loc in rep.in_location})
#     # if we fall through to here:
#     return str(value)
#
#
# def write_excel_file(protocol, serialized_noncontrol_activities, mdc: MarkdownConverter):
#     template_path = posixpath.join(os.path.dirname(os.path.realpath(__file__)),'template.xlsx')
#     wb = openpyxl.load_workbook(filename=template_path)
#     ws = wb.active  # get the default worksheet
#
#     # write header & metadata
#     ws.title = "Data Reporting"
#     ws.protection.enable()
#     ws['D1'] = protocol.name
#     ws['D1'].comment = openpyxl.comments.Comment(protocol.identity, "PAML autogeneration, do not modify", height=24,
#                                                  width=1000)
#     for row in ws['C2:C4']:
#         for cell in row: cell.protection = openpyxl.styles.Protection(locked=False)  # unlock metadata locations
#     header_style = ws['A1'].font
#     row_offset = 7  # starting point for entries
#
#     # write each value set, incrementing each time
#     value_steps = (step for step in range(len(serialized_noncontrol_activities)) if
#                    isinstance(serialized_noncontrol_activities[step], paml.Value))
#     for step in value_steps:
#         coord = 'A' + str(row_offset)
#         ws[coord] = 'Report from Step ' + str(step + 1)
#         ws[coord].font = copy(header_style)
#         value_locations = mdc.protocol_typing.flow_values[serialized_noncontrol_activities[step].input_flows().pop()]
#         block_height = excel_write_flow_value(mdc.document, value_locations, ws, row_offset + 1)
#         row_offset += block_height + 2
#
#     wb.save(protocol.display_id + '.xlsx')
#
