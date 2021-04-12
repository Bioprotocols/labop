import sbol3
import paml
import openpyxl
import tyto
import uuid
import warnings


##############################
# Set up for conversion
print('Beginning conversion of Excel to SBOL')

doc = paml.Document()

wb = openpyxl.load_workbook(filename ='../test/testfiles/BBN_LUDOX_OD_calibration_2021-3.xlsx')
ws = wb['Data Reporting']

##############################
# Pull header materials from the Excel file
print('Transforming metadata to SBOL')
# pull the ID & metadata for the protocol from the comment
protocol_id = ws['D1'].comment.text
protocol_name = ws['D1'].value
person_executing = ws['C2'].value
date_executed = ws['C3'].value
date_reported = ws['C4'].value

execution_id = str(uuid.uuid4()).replace('-','_')
execution = sbol3.Activity(protocol_id+'_Execution_'+execution_id)
execution.types.append(protocol_id) # This execution is of the protocol
doc.add(execution)

experiment = sbol3.Experiment(protocol_id+'_Results_'+execution_id)
experiment.name = protocol_name+" Execution"
experiment.description = 'Executed by '+str(person_executing)+' on '+str(date_executed)+"; reported on "+str(date_reported)
experiment.generated_by.append(execution.identity)
doc.add(experiment)

##############################
# find all cells below the header that have comments, and create an ExperimentalData, Implementation, and provenance for each
print('Transforming data to SBOL')
count = 0
for row in ws.iter_rows(min_row=6):
    for cell in row:
        if cell.comment: # any cell with a comment should have data
            count += 1
            data = cell.value
            comment_ids = cell.comment.text.split(' ')
            assert len(comment_ids)==2 # should be precisely two IDs
            implementation_id = comment_ids[0]+"_Execution_"+execution_id
            component_id = comment_ids[1]
            if data is None:
                warnings.warn('Expected value in cell '+cell.coordinate+' but found none')
            else:
                # find or make an implementation
                implementation = doc.find(implementation_id)
                if implementation is None:
                    implementation = sbol3.Implementation(implementation_id)
                    doc.add(implementation)
                    experiment.members.append(implementation)
                    implementation.derived_from.append(component_id)
                else:
                    assert implementation.derived_from[0] == component_id
                # stick the value on it as a measure
                measure = sbol3.Measure(data,tyto.OM.get_uri_by_term('number'))
                implementation.measures.append(measure)

print('Processed '+str(count)+' data entries')
print(' Note: still need to add information about what the measure is quantifying, avoid collisions when same sample is measuremed multiple times')

doc.write('BBN_LUDOX_execution_2021_3_report.json', 'json-ld')
doc.write('BBN_LUDOX_execution_2021_3_report.ttl', 'turtle')

print('writing complete')
